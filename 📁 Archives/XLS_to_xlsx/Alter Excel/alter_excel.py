import openpyxl
from xls_to_xlsx_converter import convert_xls_to_xlsx

def automatizar_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Encontra a primeira linha vazia na coluna K e copia linhas para o lado
    max_row = ws.max_row
    target_col = ws.max_column + 2  # Define a coluna de destino (espaçando para evitar sobreposição) = R até AD

    for row in range(1, max_row + 1):
        if ws.cell(row=row, column=11).value is None:  # Coluna K é a 11ª
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=target_col + col - 1).value = ws.cell(row=row, column=col).value

    # Deletar células com '#' até 'Modalidade' + 7 linhas acima
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == '#':
                start_row = cell.row
                end_row = max(1, start_row - 7)
                ws.delete_rows(end_row, start_row - end_row + 1)
                break

    # Deletar células com '* Parcelas Alteradas' + 2 linhas abaixo
    for row in ws.iter_rows(min_col=15, max_col=15):
        for cell in row:
            if cell.value == '* Parcelas Alteradas':
                start_row = cell.row
                ws.delete_rows(start_row, 3)  # Deleta a linha atual e mais 2 abaixo
                break

    # Verificar se os números na coluna B estão em ordem crescente
    numbers_in_order = True
    previous_value = None
    for row in ws.iter_rows(min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Verifica se é um número
                if previous_value is not None and cell.value <= previous_value:
                    numbers_in_order = False
                    break
                previous_value = cell.value

    if not numbers_in_order:
        print("Os números na coluna B não estão em ordem crescente.")
    else:
        print("Os números na coluna B estão em ordem crescente.")

    wb.save(file_path)

# Entrada do nome do arquivo
nome_arquivo = input("Digite o nome do arquivo, sem o final do arquivo: ")

arquivo_XLS = nome_arquivo + ".XLS"
arquivo_xlsx = nome_arquivo + ".xlsx"

if arquivo_XLS:
    convert_xls_to_xlsx(arquivo_XLS, arquivo_xlsx)
    automatizar_excel(arquivo_xlsx)
